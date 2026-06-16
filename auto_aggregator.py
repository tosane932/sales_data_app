import os
import openpyxl
import config  # 一元管理された設定情報をインポート

# 集計対象となるExcelファイルが格納されているフォルダパスをconfigから取得
folder_path = config.PAST_FOLDER


def get_filtered_sales_data(target_year=None, target_month=None):
    """
    指定された年(target_year)や月(target_month)で売上ファイルを絞り込んで集計する関数。
    引数がNoneの場合は、過去の全期間を対象に集計を行う。
    """
    total_sales = {}

    # 対象フォルダが存在しない場合は、エラーを防ぐため空の辞書を返して即座に終了する
    if not os.path.exists(folder_path):
        print(f"DEBUG: フォルダが見つかりません -> {folder_path}")
        return total_sales

    files = os.listdir(folder_path)
    print(f"DEBUG: フォルダ内の全ファイル -> {files}")

    for file_name in files:
        # アプリが生成した売上Excelファイル（例: 売上高202606.xlsx）のみを処理対象とする
        if not file_name.startswith("売上高") or not file_name.endswith(".xlsx"):
            continue

        try:
            # ファイル名から年月を表す文字列（202606など）を抽出し、年と月に分解して整数化
            date_str = file_name.replace("売上高", "").replace(".xlsx", "")
            year = int(date_str[:4])
            month = int(date_str[4:6])
        except Exception as e:
            # ファイル名の形式不正などで解析に失敗した場合は、ログを残して次のファイルを処理する
            print(f"DEBUG: ファイル名解析エラー -> {file_name} ({e})")
            continue

        # 指定された絞り込み条件（引数）に合致しない期間のファイルはスキップする
        if target_year and year != target_year:
            continue
        if target_month and month != target_month:
            continue

        print(f"DEBUG: 🔵集計開始 -> {file_name}")

        file_full_path = os.path.join(folder_path, file_name)
        # 数式そのものではなく、計算された後の数値を取得するために data_only=True を明示
        wb = openpyxl.load_workbook(file_full_path, data_only=True)
        ws = wb.active

        # Excelの構造が「価格・数量・合計金額」の3列1セットになっているため、3列飛ばしでループを回す
        for col_idx in range(3, ws.max_column + 1, 3):
            prod_name = ws.cell(row=1, column=col_idx).value

            if prod_name:
                for row_idx in range(3, ws.max_row + 1):
                    # 日付列（1列目）の文字列が「総合計」に達した時点で、該当シートの縦方向のデータ走査を打ち切る
                    cell_val = ws.cell(row=row_idx, column=1).value
                    if str(cell_val) == "総合計":
                        break

                    qty = ws.cell(row=row_idx, column=col_idx + 1).value

                    # セルが空文字（None）や未入力でないことを確認し、安全に型変換して集計に加算する
                    if qty is not None and str(qty).strip() != "":
                        try:
                            # 文字列としての小数（"10.0"など）が直接intに変換できないエラーを防ぐため、一度floatを経由させる
                            total_sales[prod_name] = total_sales.get(prod_name, 0) + int(float(qty))
                        except ValueError:
                            # 店長さんが「休」などの文字列を数量セルに入力していた場合、エラーにせず安全にスキップする
                            pass

    print(f"DEBUG: 🟢最終集計結果 -> {total_sales}")
    return total_sales
