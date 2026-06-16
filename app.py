import os
import datetime
import calendar
from flask import Flask, render_template, request, send_file, jsonify
import openpyxl
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection
from auto_aggregator import get_filtered_sales_data

# 最新の Google GenAI SDK クライアントモジュールを導入
from google import genai
from google.genai import types

# 一元管理された設定情報をインポート
import config

app = Flask(__name__)

# アプリ起動時に保存先フォルダが存在しない場合は自動生成する
if not os.path.exists(config.PAST_FOLDER):
    os.makedirs(config.PAST_FOLDER)


def generate_excel(year, month, products_data):
    """
    ユーザーがWebから入力した商品データを基に、
    誤操作を防ぐシート保護を施したオーダーメイドのExcelファイルを生成する。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月間売上入力"
    ws.protection.enabled = True  # ユーザーの数式破壊を防ぐためシート全体を保護

    # 表のデザイン（罫線・背景色）を定義
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    border_header_normal = Border(left=thin, right=thin, top=medium, bottom=medium)
    border_header_block_end = Border(left=thin, right=medium, top=medium, bottom=medium)
    border_data_normal = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_data_block_end = Border(left=thin, right=medium, top=thin, bottom=thin)

    fill_product = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    fill_item = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_total = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # 視認性向上のため、土日の行に適用する背景色
    weekday_colors = {
        "日": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "土": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    }

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # 固定ヘッダー（日付・曜日列）の設定
    ws["A1"] = "日付"
    ws["B1"] = "曜日"
    ws["A2"] = "固定"
    ws["B2"] = "自動算出"
    for cell_id in ["A1", "B1", "A2", "B2"]:
        ws[cell_id].alignment = align_center

    # フォームから受け取った商品情報をもとに、3列1ブロック（価格・数量・合計）のヘッダーを構築
    current_col = 3
    for prod in products_data:
        c_price = openpyxl.utils.get_column_letter(current_col)
        c_qty = openpyxl.utils.get_column_letter(current_col + 1)
        c_amount = openpyxl.utils.get_column_letter(current_col + 2)

        ws[f"{c_price}1"] = prod["name"]
        ws.merge_cells(f"{c_price}1:{c_amount}1")
        ws[f"{c_price}1"].fill = fill_product
        ws[f"{c_price}1"].font = Font(name="Noto Sans CJK SC", size=11, bold=True)
        ws[f"{c_price}1"].alignment = align_center

        ws[f"{c_price}2"] = "価格"
        ws[f"{c_qty}2"] = "数量"
        ws[f"{c_amount}2"] = "合計金額"

        for col_let in [c_price, c_qty, c_amount]:
            ws[f"{col_let}2"].fill = fill_item
            ws[f"{col_let}2"].alignment = align_center

        current_col += 3

    # 指定された年月の暦に応じて、日付行と連動する数式を1日ずつ生成
    _, max_days = calendar.monthrange(year, month)
    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]

    for day in range(1, max_days + 1):
        row_idx = 2 + day
        date_val = datetime.date(year, month, day)

        ws[f"A{row_idx}"] = date_val
        ws[f"A{row_idx}"].number_format = 'm/d'
        ws[f"A{row_idx}"].border = border_data_normal
        ws[f"A{row_idx}"].alignment = align_center

        ja_w = weekday_ja[date_val.weekday()]
        ws[f"B{row_idx}"] = ja_w
        ws[f"B{row_idx}"].border = Border(left=thin, right=medium, top=thin, bottom=thin)
        ws[f"B{row_idx}"].alignment = align_center
        if ja_w in weekday_colors:
            ws[f"B{row_idx}"].fill = weekday_colors[ja_w]

        p_col = 3
        for prod in products_data:
            c_p = openpyxl.utils.get_column_letter(p_col)
            c_q = openpyxl.utils.get_column_letter(p_col + 1)
            c_a = openpyxl.utils.get_column_letter(p_col + 2)

            # マスタとなる3行目だけ数値を書き込み、4行目以降は絶対参照にすることでファイルサイズと冗長性を削減
            if row_idx == 3:
                ws[f"{c_p}3"] = prod["price"]
            else:
                ws[f"{c_p}{row_idx}"] = f"=${c_p}$3"

            # 各日の売上金額を計算する数式（合計金額 = 価格 * 数量）
            ws[f"{c_a}{row_idx}"] = f"={c_p}{row_idx}*{c_q}{row_idx}"
            p_col += 3

    # 月末行の直下に「総合計」行を作成し、縦方向のSUM関数を設定
    total_row = 2 + max_days + 1
    ws[f"A{total_row}"] = "総合計"
    ws[f"A{total_row}"].font = Font(name="Noto Sans CJK SC", size=10, bold=True)
    ws[f"A{total_row}"].fill = fill_total
    ws[f"A{total_row}"].alignment = align_center
    ws[f"A{total_row}"].border = border_data_normal
    ws[f"B{total_row}"].fill = fill_total
    ws[f"B{total_row}"].border = Border(left=thin, right=medium, top=thin, bottom=thin)

    p_col = 3
    max_col_idx = 2 + (len(products_data) * 3)

    for _ in products_data:
        c_p = openpyxl.utils.get_column_letter(p_col)
        c_q = openpyxl.utils.get_column_letter(p_col + 1)
        c_a = openpyxl.utils.get_column_letter(p_col + 2)

        ws[f"{c_p}{total_row}"] = ""
        ws[f"{c_q}{total_row}"] = f"=SUM({c_q}3:{c_q}{total_row-1})"
        ws[f"{c_a}{total_row}"] = f"=SUM({c_a}3:{c_a}{total_row-1})"

        for col_let in [c_p, c_q, c_a]:
            cell = ws[f"{col_let}{total_row}"]
            cell.fill = fill_total
            cell.font = Font(name="Noto Sans CJK SC", bold=True)
            cell.alignment = align_right
        p_col += 3

    # マイナス値の入力を防ぐデータバリデーション（入力規則）を設定
    dv_qty = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
    dv_qty.promptTitle = "【数量の入力】"
    dv_qty.prompt = "本日の販売個数を入力してください。"
    dv_qty.showInputMessage = True
    ws.add_data_validation(dv_qty)

    # ユーザーが編集可能なセル（数量列のみ）のロックを解除
    for r in range(3, total_row + 1):
        for c in range(3, max_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            mod = (c - 3) % 3

            if mod == 2:
                cell.border = border_data_block_end
                cell.alignment = align_right
            else:
                cell.border = border_data_normal
                cell.alignment = align_right
                # 総合計行を除く、数量入力列（mod == 1）のみ保護を外して入力規則の対象とする
                if r != total_row and mod == 1:
                    cell.protection = Protection(locked=False)
                    dv_qty.add(cell)

    # 見映えを整えるため、ヘッダー（1〜2行目）の太枠・細枠の罫線を最終調整
    for r in [1, 2]:
        for c in range(1, max_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
            elif c == 2:
                cell.border = Border(left=thin, right=medium, top=medium, bottom=medium)
            elif (c - 3) % 3 == 2:
                cell.border = border_header_block_end
            else:
                cell.border = border_header_normal

    # データの文字数に合わせた各列幅の最適化
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    for c in range(3, max_col_idx + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        mod = (c - 3) % 3
        if mod == 0:
            ws.column_dimensions[col_letter].width = 9
        elif mod == 1:
            ws.column_dimensions[col_letter].width = 9
        else:
            ws.column_dimensions[col_letter].width = 15

    # 定義されたファイル命名規則に則ってconfig内で定義されたフォルダへ保存
    filename = f"売上高{year}{month:02d}.xlsx"
    file_path = os.path.join(config.PAST_FOLDER, filename)
    wb.save(file_path)
    return file_path, filename


def _generate_ai_advice(ranked_sales):
    """
    売上データを解析し、Gemini APIを用いて経営アドバイスのテキストを生成する共通関数。
    通常表示と非同期通信（API）の両方から呼び出され、エラーハンドリングを一元化している。
    """
    if not ranked_sales:
        return "売上データがまだないため、アドバイスを生成できません。"

    try:
        api_key = os.environ.get("GEMINI_API_KEY")
        # デプロイ環境を考慮し、特定のOS名に依存しない汎用的なエラーメッセージに修正
        if not api_key:
            return "🚨【設定未完了】環境変数に GEMINI_API_KEY が登録されていません。ダッシュボードの設定を確認してください。"

        # GenAIクライアントの初期化およびコンサルタントとしてのプロンプト構築
        client = genai.Client(api_key=api_key)
        sales_summary = ", ".join([f"{name}: {qty}個" for name, qty in ranked_sales])

        prompt = f"""
        あなたは街の優しいベーカリーの優秀な経営コンサルタントです。
        以下の売上データ（商品名と販売数量）を見て、店長さんが明日から元気にお店を経営できるような、
        温かくて具体的なアドバイスを3文以内で親しみやすく教えてください。
        
        データ: {sales_summary}
        """

        # configで一元管理しているモデル名を使用してアドバイスをリクエスト
        response = client.models.generate_content(
            model=config.GEMINI_MODEL,
            contents=prompt,
        )
        return response.text

    except Exception as e:
        # 短時間での過剰なリクエストによりAPIの速度制限（429エラー）に到達した場合の処理
        if "429" in str(e) or "RESOURCE_EXHAUSTED" in str(e):
            return "☕【AIが少し休憩中です】\n短時間にたくさんデータを抽出したため、AIの速度制限がかかっております。お手数ですが、10秒〜20秒ほどあけて、もう一度「データを抽出」ボタンを押してみてください。"
        return f"AIアドバイスを生成中に一時的なエラーが発生しました。（デバッグ用: {e}）"


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        year = int(request.form.get("year"))
        month = int(request.form.get("month"))

        product_names = request.form.getlist("prod_name")
        product_prices = request.form.getlist("prod_price")

        products_data = []
        for name, price in zip(product_names, product_prices):
            if name.strip():
                products_data.append({
                    "name": name.strip(),
                    "price": int(price) if price.isdigit() else 0
                })

        # 指定の条件で保護されたExcelファイルを生成して保存
        generate_excel(year, month, products_data)

        # スマホ環境での操作性を担保するため、ファイル送信ではなく完了画面をレンダリングして返却
        return render_template("success.html", year=year, month=month)

    return render_template("index.html")


@app.route("/dashboard")
def dashboard():
    """
    通常の画面遷移でダッシュボードを表示する際のルーティング。
    初期表示時、またはクエリパラメータによる期間絞り込み時に動作する。
    """
    year_param = request.args.get("year")
    month_param = request.args.get("month")

    target_year = int(year_param) if year_param else None
    target_month = int(month_param) if month_param else None

    # 保存先フォルダから該当するデータを抽出・集計
    sales_data = get_filtered_sales_data(target_year, target_month)
    ranked_sales = sorted(sales_data.items(), key=lambda item: item[1], reverse=True)

    chart_labels = [name for name, qty in ranked_sales]
    chart_values = [qty for name, qty in ranked_sales]

    # 共通化した関数を呼び出してAIアドバイスを取得
    ai_advice = _generate_ai_advice(ranked_sales)

    return render_template("dashboard.html", 
                           sales=sales_data, 
                           ranked_sales=ranked_sales,
                           chart_labels=chart_labels,
                           chart_values=chart_values,
                           ai_advice=ai_advice,
                           year=target_year or "全期間",
                           month=target_month or "全期間")


@app.route("/api/dashboard-data")
def api_dashboard_data():
    """
    画面をリロードせずにデータを非同期（Ajax）で更新するためのエンドポイント。
    """
    year_param = request.args.get("year")
    month_param = request.args.get("month")

    target_year = int(year_param) if year_param else None
    target_month = int(month_param) if month_param else None

    sales_data = get_filtered_sales_data(target_year, target_month)
    ranked_sales = sorted(sales_data.items(), key=lambda item: item[1], reverse=True)

    chart_labels = [name for name, qty in ranked_sales]
    chart_values = [qty for name, qty in ranked_sales]

    # 共通化した関数から安全にAIアドバイスを取得
    ai_advice = _generate_ai_advice(ranked_sales)

    # 呼び出し元のJavaScriptが処理しやすいようにJSON形式に整形して返却
    return jsonify({
        "ranked_sales": ranked_sales,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "ai_advice": ai_advice,
        "period_text": f"{target_month}月度" if target_month else "全期間"
    })


if __name__ == "__main__":
    # 外部接続を許可し、デバッグモードを有効にしてアプリケーションをローカル起動
    app.run(debug=True, host="0.0.0.0", port=5000)
