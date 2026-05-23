import os
import datetime
import calendar  # 🌟 月の日数を自動判定するための救世主
import openpyxl
from openpyxl.styles import Font, Side, Border, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.protection import Protection

def build_sales_sheet(year, month, product_list):
    """
    指定された「年」「月」「商品リスト」に基づいて、
    日数を自動判定して完全に適合する売上入力シートを動的生成する関数
    """
    past_folder = "過去売上高"
    if not os.path.exists(past_folder):
        os.makedirs(past_folder)

    wb = openpyxl.Workbook()
    ws = wb.active
    # シート名に対象の年月を刻みます（例：2026年7月売上入力）
    ws.title = f"{year}年{month}月売上入力"

    # シート保護を有効化
    ws.protection.enabled = True

    # 罫線・デザインの定義
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    border_header_normal = Border(left=thin, right=thin, top=medium, bottom=medium)
    border_header_block_end = Border(left=thin, right=medium, top=medium, bottom=medium)
    border_data_normal = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_data_block_end = Border(left=thin, right=medium, top=thin, bottom=thin)

    fill_product = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
    fill_item = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_total = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    weekday_colors = {
        "日": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "月": PatternFill(start_color="FFD9E8", end_color="FFD9E8", fill_type="solid"),
        "火": PatternFill(start_color="FFEBCC", end_color="FFEBCC", fill_type="solid"),
        "水": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "木": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "金": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "土": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    }

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # 基本見出し
    ws["A1"] = "日付"
    ws["B1"] = "曜日"
    ws["A2"] = "固定"
    ws["B2"] = "自動算出"
    for cell_id in ["A1", "B1", "A2", "B2"]:
        ws[cell_id].alignment = align_center

    # 商品列の動的構築
    current_col = 3
    for prod in product_list:
        c_price = openpyxl.utils.get_column_letter(current_col)
        c_qty = openpyxl.utils.get_column_letter(current_col + 1)
        c_amount = openpyxl.utils.get_column_letter(current_col + 2)
        
        ws[f"{c_price}1"] = prod
        ws.merge_cells(f"{c_price}1:{c_amount}1")
        ws[f"{c_price}1"].fill = fill_product
        ws[f"{c_price}1"].font = Font(name="Noto Sans CJK SC", size=11, bold=True)
        ws[f"{c_price}1"].alignment = align_center
        
        ws[f"{c_price}2"] = "価格"
        ws[f"{c_qty}2"] = "数量"
        ws[f"{c_amount}2"] = "合計金額"
        
        for col_let in [c_price, c_qty, c_amount]:
            ws[f"{col_let}2"].fill = fill_item
            ws[f"{col_let}2"].alignment = align_center
        
        current_col += 3

    # 🌟【超重要】指定された年月の「実際の日数」を自動判定！
    # calendar.monthrange(year, month) は (最初の日の曜日, その月の日数) を返します
    _, max_days = calendar.monthrange(year, month)

    weekday_ja = ["月", "火", "水", "木", "金", "土", "日"]
    
    # 割り出された実際の日数分（1日〜max_days日）だけループを回す
    for day in range(1, max_days + 1):
        row_idx = 2 + day
        date_val = datetime.date(year, month, day)
        
        # A列：日付
        ws[f"A{row_idx}"] = date_val
        ws[f"A{row_idx}"].number_format = 'm/d'
        ws[f"A{row_idx}"].border = border_data_normal
        ws[f"A{row_idx}"].alignment = align_center
        
        # B列：曜日（2026年7月の日付に合った正しい曜日が100%自動で入ります）
        ja_w = weekday_ja[date_val.weekday()]
        ws[f"B{row_idx}"] = ja_w
        ws[f"B{row_idx}"].border = Border(left=thin, right=medium, top=thin, bottom=thin)
        ws[f"B{row_idx}"].alignment = align_center
        if ja_w in weekday_colors:
            ws[f"B{row_idx}"].fill = weekday_colors[ja_w]
        
        # 商品ごとの数式割り当て（価格連動 ＆ 掛け算）
        p_col = 3
        for _ in product_list:
            c_p = openpyxl.utils.get_column_letter(p_col)
            c_q = openpyxl.utils.get_column_letter(p_col + 1)
            c_a = openpyxl.utils.get_column_letter(p_col + 2)
            
            if row_idx > 3:
                ws[f"{c_p}{row_idx}"] = f"=${c_p}$3"
            
            ws[f"{c_a}{row_idx}"] = f"={c_p}{row_idx}*{c_q}{row_idx}"
            p_col += 3

    # 🌟【大進化】総合計行の位置も「日数の真下」へ動的にスライド！
    total_row = 2 + max_days + 1  # 例：31日ある月なら、2 + 31 + 1 ＝ 34行目が総合計になる
    
    ws[f"A{total_row}"] = "総合計"
    ws[f"A{total_row}"].font = Font(name="Noto Sans CJK SC", size=10, bold=True)
    ws[f"A{total_row}"].fill = fill_total
    ws[f"A{total_row}"].alignment = align_center
    ws[f"A{total_row}"].border = border_data_normal

    ws[f"B{total_row}"] = "" 
    ws[f"B{total_row}"].fill = fill_total
    ws[f"B{total_row}"].border = Border(left=thin, right=medium, top=thin, bottom=thin)

    p_col = 3
    max_col_idx = 2 + (len(product_list) * 3)

    for _ in product_list:
        c_p = openpyxl.utils.get_column_letter(p_col)
        c_q = openpyxl.utils.get_column_letter(p_col + 1)
        c_a = openpyxl.utils.get_column_letter(p_col + 2)
        
        ws[f"{c_p}{total_row}"] = ""
        # 🌟 SUMの範囲も、その月の最終行（total_row - 1）まで自動可変対応！
        ws[f"{c_q}{total_row}"] = f"=SUM({c_q}3:{c_q}{total_row-1})"
        ws[f"{c_a}{total_row}"] = f"=SUM({c_a}3:{c_a}{total_row-1})"
        
        for col_let in [c_p, c_q, c_a]:
            cell = ws[f"{col_let}{total_row}"]
            cell.fill = fill_total
            cell.font = Font(name="Noto Sans CJK SC", bold=True)
            cell.alignment = align_right
        p_col += 3

    # 入力規則
    dv_price = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
    dv_price.promptTitle = "【価格の入力】"
    dv_price.prompt = "初日の単価（円）を入力してください。自動的に下行へ連動します。"
    dv_price.showInputMessage = True
    ws.add_data_validation(dv_price)

    # 数量の最大制限をかけていない安全仕様
    dv_qty = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0")
    dv_qty.promptTitle = "【数量の入力】"
    dv_qty.prompt = "本日の販売個数を入力してください。"
    dv_qty.showInputMessage = True
    ws.add_data_validation(dv_qty)

    # 動的な最終行（total_row）に合わせたセルロック制御
    for r in range(3, total_row + 1):
        for c in range(3, max_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            mod = (c - 3) % 3
            
            if mod == 2:
                cell.border = border_data_block_end
                cell.alignment = align_right
            else:
                cell.border = border_data_normal
                cell.alignment = align_right
                
                if r != total_row:
                    if mod == 0 and r == 3:
                        cell.protection = Protection(locked=False) # 初日価格のみ開放
                        dv_price.add(cell)
                    elif mod == 1:
                        cell.protection = Protection(locked=False) # 数量は毎日開放
                        dv_qty.add(cell)

    # 見出し行枠線調整
    for r in [1, 2]:
        for c in range(1, max_col_idx + 1):
            cell = ws.cell(row=r, column=c)
            if c == 1:
                cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
            elif c == 2:
                cell.border = Border(left=thin, right=medium, top=medium, bottom=medium)
            elif (c - 3) % 3 == 2:
                cell.border = border_header_block_end
            else:
                cell.border = border_header_normal

    # 列幅最適化
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 6
    for c in range(3, max_col_idx + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        mod = (c - 3) % 3
        if mod == 0:
            ws.column_dimensions[col_letter].width = 9
        elif mod == 1:
            ws.column_dimensions[col_letter].width = 9
        else:
            ws.column_dimensions[col_letter].width = 15

    # 保存ファイル名に年月を組み込み、上書き事故を防ぐ
    file_name = f"売上高入力シート_{year}_{month:02d}.xlsx"
    wb.save(file_name)
    print(f"【究極汎用化】{year}年{month}月（実日数：{max_days}日）のシート『{file_name}』を完璧に自動生成しました！")

if __name__ == "__main__":
    # 🌟 マスター、ここを自由に変えるだけで何年何月でも一撃生成されます！
    # 実験として、ご要望の「2026年 7月（31日まである月）」を指令します。
    target_year = 2026
    target_month = 7
    
    current_menu = ["食パン", "メロンパン", "クロワッサン", "コロネ", "ホットドッグ", "カレーパン"]
    build_sales_sheet(target_year, target_month, current_menu)